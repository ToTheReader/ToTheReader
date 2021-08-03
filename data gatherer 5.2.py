# runs through EEBO-TCP XML corpus directory
# identifies desired texts by lists of div types
# extracts required data of desired texts
# writes values to an excel file
# outputs xml files of each text

# new in version 3:
# Analysis of preface and assignation to dict/excel now in the function 'describe()' (c. ll.28-118)
# Types listed in 'pluraltypes' (e.g. encomia, define din . 29) are unpacked and treated as separate prefaces (e.g. encomium) (c. ll. 172-196, 231-248)
# Form recognition (c. ll. 99-108) now based on presence of <P>, not <L>
# 'parts' removed from excludedtypes
# tabletypes introduced (just errata at the moment), and loop created to filter out divs that contain only tables (c. ll. 207-220, ll. 273-286)

# new in version 4:
# describe()s relevant BODY DIV1 and DIV2 types (defined by lists bodytypes and body_div2_types)
# describe()s relevant DIV2 types within excluded DIV1 types for FRONT and BACK (as defined by list, front_and_back_div2_types)
# 3rd 'form' option added to describe(): 'Prose/Verse', for texts containing both P and L tags
# checks whether relevant DIV2s are just 'letters' in index before describing DIV2

#new in version 5:
# new format for ID numbers: ESTCno_Section_Div1no_Div2no
# populates edition table with ESTC no. and preface IDS (for ones with 0 prefaces too)
# saves div as separate xml file
# new lists made (one for each section) and made inclusive, not exclusive
# removed dictionary data structure that was being compiled parallel to excel output

# 5.1 has some restructuring and edits to improve speed.
# removed some redundant parsing, e.g.:
# 
#        preface_content = bs(str(div), 'xml')
#        div2s = preface_content.find_all('DIV2')

# now

#        div2s = div.preface_content.find_all('DIV2)')

# as div is already readable by bs. Same applies to beginning of 'describe' function

# streamlined the describe function (original structure was intended to share values between dict and excel output, which is now redundant without the dictionary)
# excel is now saved only once at end of the loop (i.e. once for each preface entry)

# 5.2
# an if clause added to check that the file ends in .xml (to avoid .DS_Store and other file types)
# Adding pathlib to make compatible with Windows and Mac

import bs4
import lxml
from bs4 import BeautifulSoup as bs
import os
import openpyxl
import time
from pathlib import Path

start_time = time.time()

# Set directory here
directory = Path('Folder', 'subfolder if there is one')

os.chdir(str(directory))

# DIV TYPE LISTS
# FRONT
FRONT_DIV1S = ['dedication',
 'to the reader',
 'poem',
 'encomium',
 'preface',
 'errata',
 'prologue',
 'letter',
 'argument',
 'author to the reader',
 'epigraph',
 'introduction',
 'translator to the reader',
 'printer to the reader',
 'summary',
 'encomia',
 'prayer',
 'to the author',
 'sonnet',
 'notice',
 'note',
 'poem to the reader',
 "translator's dedication",
 'text',
 'author to book',
 "author's dedication",
 'epigram',
 'epitaph',
 'dedicatory poem',
 'biography',
 'prefatory letter',
 "translator's preface",
 'preface to the reader',
 "author's note",
 'poems',
 'proem',
 'acrostic',
 'publisher to the reader',
 'acrostic poem',
 'dialogue',
 'postscript',
 "author's preface",
 'apology',
 'elegy',
 'book to the reader',
 'to the translator',
 'commentary',
 'answer',
 'account',
 'glossary',
 'admonition',
 'abstract',
 'anagram',
 'front matter',
 "author's prologue",
 'preamble',
 'epilogue',
 'exhortation',
 'prefatory poem',
 'to the Christian reader',
 'summaries',
 'declaration',
 'to the courteous reader',
 "translator's prologue",
 'to the book',
 'dedicatory letter',
 'opinion',
 'book to author',
 'dedicatory verse or poem',
 "Caxton's preface",
 'list of reasons to sing',
 'explanation of frontispiece',
 'note to the reader',
 'confession',
 'proposition',
 'panegyric',
 'attestation',
 'commendation',
 "publisher's advertisement",
 'instructions',
 "translator's note",
 'reply',
 'translator to book',
 'author to printer',
 "author's letter",
 'stationer to the reader',
 "printer's note",
 'letter to author',
 'dedications',
 'distichs',
 'rule',
 'discourse',
 'praise',
 'translator to author',
 'invocation',
 'dedicatory epistle',
 'instruction',
 'notes',
 'satire',
 'verse dedication',
 'introduction to Welsh',
 'addendum',
 'ode',
 'challenge',
 'author to his book',
 "publisher's note",
 'to the Catholic reader',
 'prophecy',
 'note on the text',
 'response to preface',
 'sonnet to the reader',
 'introductory poem',
 'definition',
 'summary of Colossians',
 'response',
 'to printer',
 'advertisement',
 'to the traveller',
 'author to colleagues',
 'memorial',
 'poem on frontispiece',
 'abstract of part',
 'address',
 'addenda',
 'author to the printer',
 'editor to the reader',
 'proof',
 'induction',
 'author to friend',
 "author's preface to the reader",
 'poem to book',
 'to the printer',
 'prologue to the reader',
 'author to critic',
 'printer to author',
 'bookseller to the reader',
 'envoy',
 'preface to distillers',
 'epigrams',
 'translation',
 'premonition',
 "translator's notes",
 'by the same author',
 'to the censors',
 'prelude',
 'to poets',
 'commission',
 'acrostic encomium',
 'epithalamium',
 'author to guild',
 'to the Catholic-like Protestant reader',
 'counsel to the reader',
 'summary of argument',
 'last words',
 'notice to the reader',
 "Cranmer's prologue",
 'poem to the censors',
 'annotator to the reader',
 'introductory letter',
 "author's reply to encomia",
 'instruction to the reader',
 'lectionary instructions',
 'introductory encomium',
 'to husbandmen',
 "publisher's dedication",
 'deciation',
 'instruction to printer',
 'compositor to the reader',
 'Caxton to the reader',
 'epistle dedicatory',
 'rules of French pronunciation',
 'prognosticon',
 "Latin translator's note",
 'list of principal positions',
 'preface to parsons and curates in the diocese of London',
 'exposition',
 "author's friend to author's rival",
 'prose summary',
 'to hostile courtiers',
 'considerations',
 'to his book',
 "author's invitation",
 'advertisement to the reader',
 'author to his lady',
 'engraved title page with verse explanation',
 'St Bernard to abbots',
 'acrostic to the reader',
 'epistle',
 'epigraph to the reader',
 "printer's preface",
 'translator to the virgin reader',
 'explanation of analysis',
 "translator's preface to the author",
 'sanction',
 'author to his muse',
 'to protestant reader',
 "printer's dedication",
 'author to censor',
 'the author to the reader',
 'from the printer',
 'to ladies',
 'directions for distance tables',
 'directions for use',
 'royal notice',
 'commentary on prologue',
 'author to surgeons and physicians',
 'instructions in Scripture reading',
 'lists of ingredients',
 'letter to the interlocutor',
 'miscellaneous features',
 'prefatory poems',
 'reminder of crucifixion',
 'acrostic maxims',
 'symbol',
 'overview',
 'medical miscellany',
 'another preface to the reader',
 'meaning of picture',
 'to Protestant readers',
 'author to William Alablaster',
 'summary of gospel of John',
 'garland',
 'translator to friend',
 'list of words',
 "editor's dedication",
 'notice about conversion of measures',
 'distich',
 'verse',
 'errata note',
 'title page emblem explained',
 'author to priests',
 'key to title page',
 "publisher's notice",
 'author to his students',
 'summary of contradictions',
 'to the Papist reader',
 'to the male reader',
 'to the English nation',
 'letters of credence',
 "translator's foreword",
 'preface to fourth treatise',
 'translator to the author',
 'to physicians',
 "to the author's nephew",
 'prefatory verse or poem',
 'afterword to the reader',
 'examination',
 'author biography',
 'engraved dedication',
 'to the citizens of London',
 'Morphorius to Pasquill',
 'to virtuous reader',
 'to his opponent',
 'to the satirist',
 'to muse',
 'prefatory material',
 "King's note to author",
 'to the English reader',
 'to the Italian reader',
 'to the planters',
 'exposition of geometrical terms',
 'royal injunction',
 'to his muse',
 'Morphorius to the reader',
 'to the Christian Reader',
 'author to carpers',
 'to the readers',
 'verse prayer',
 'in place of encomium',
 'reason for writing',
 'dedication to the reader',
 'brief poems',
 'to Catholic readers',
 'explanation of titlepage',
 'to the buyer',
 'to non-English-speaking readers',
 'commentary on the dedication',
 'to kindly courtiers',
 'further to the reader',
 'acrostic meditations',
 'introduction to treatise of arithmetic',
 'to performer',
 'to God',
 'mock license',
 'dedicatory ode',
 'rules of admission',
 'biographical introduction',
 'history of poem',
 'instructions for study',
 'translator to readers',
 'meditation and confession',
 'general prologue',
 'explanation of names',
 'description of Cyprus',
 'paraphrase',
 'explanation of psalm',
 "translator's envoy",
 'printer to translator',
 'key to sigla',
 'acrostic dedicatory poem',
 'points under consideration',
 'answer to the preface',
 'prologue for court',
 'admonition to the reader',
 'Bishop to the reader',
 'salutation',
 'the book to the reader',
 'note on dating system',
 'speech',
 'to worthy soldiers',
 'historical summary',
 'master to author',
 'description of siege tower',
 'commentary on tables',
 "bookseller's dedication",
 'to the British reader',
 'to the Ministers',
 'To the reader',
 'to the Papists',
 'testimonies',
 "author's oath",
 'response to letter',
 'instructions for commissioners',
 'dialogue introduction',
 'to impecunious',
 'notice about abbreviations',
 'publisher to author',
 'translator to English king, James I',
 'verse preface',
 'textual note',
 'summary of interpretation',
 'instructions to the reader',
 'general description',
 "printer's apology",
 'dedicatory acrostic poem',
 'criticism',
 'to the Protestant reader',
 'complaint',
 'advert',
 'summary of reasons',
 'preface of original author',
 'to women',
 'author to British gentlemen',
 'note on ceremonies',
 "author's foreword",
 'to censor',
 'explanation of title',
 "translator's complaint",
 'to book',
 'summary of Synoptic gospels',
 'corrector to the reader',
 'explanation of terms',
 'to the non-reader',
 'to the misinterpreting reader',
 'eclogue',
 "Italian translator's note to the reader",
 "St. Paul's cathedral speaks",
 'explanation of argument',
 'to the blind reader',
 "translator's poem to the reader",
 'mnemonic poem',
 'to catholic reader',
 'introduction to the reader',
 'supplement',
 'author to cousin',
 'contents in verse',
 'to the gardener',
 'note from the author',
 'a Donatist analogy',
 'notice regarding errata',
 'note on sources',
 'couplet',
 'to the pespicuous reader',
 'author to youth',
 'to the faithful',
 'author to French king, Louis XIII',
 'historical note',
 'supplication',
 'translator to poets',
 'profession of faith poem',
 'chief points',
 "editor's preface",
 'poem and prayer',
 'life',
 'to surgeons',
 'guide to lotteries',
 'prologue to poem',
 'annotations upon preface',
 'letter of commendation',
 'anti-dedication',
 'translation of poem',
 'biographies',
 "author's statement of purpose",
 'testimonials',
 'clarification',
 'life of Lucan',
 'to the Duke of York',
 "to the reader's friend",
 'to the young Duke of York',
 'author to prelates',
 'request',
 'dedication cartouche',
 'explanation of the emblem',
 'history',
 'note about translations',
 "translator's letter",
 'acrostic poem to the reader',
 'letter to his son',
 'author to king',
 'editor to translator',
 'poem to readers',
 'verse argument',
 'challenges',
 'to carping readers',
 'hand-written dedication',
 "poetical exposition of book's contents",
 'dedicatory acrostic',
 'translator of French version to the reader',
 'verdict',
 'book',
 'author to nemesis',
 'translator to English Catholics',
 'Pasquill to Morphorius',
 'in praise of Roger Ascham',
 'poem author to book',
 'answer to dedication',
 'book to readers',
 'dedicatory material',
 'to the gentle reader',
 'letter to the reader',
 'to the Romish reader',
 'author to publisher',
 "author's postscript",
 'note on sigla',
 'objection',
 'note on origin of work',
 'prefatory epistle',
 'lamentation',
 'corrections',
 'description of triumphal procession',
 'summary poem',
 'guide to dedications',
 'Lyly to author',
 'biography of author',
 'preface to King',
 'summary of treatise',
 'summary of doctrine',
 "author's apology",
 'note on pagination',
 'to Prince Charles',
 'hymn and prayer',
 'foreword',
 'advert for game',
 'customs rates',
 'epistle to the reader',
 'author to the readers',
 'schedule',
 'printer to book',
 'petition',
 'introductory preambles',
 'defense of epitaph',
 'to papists',
 'standards of the faith',
 'holograph dedication',
 'letters',
 'dedicatory tetrastich',
 'printers to readers',
 'biblical epitome',
 'explanation of format',
 'note to printer',
 'interpreter to the reader',
 'tetrastich',
 'poem to Archpapist',
 'advert for geometrical instruments',
 'computation',
 'answer to Mr Smythe',
 'to his parishioners',
 'acrositc dedication',
 'verse epistle',
 'summary of work',
 'author to christian reader',
 'instructions for reading',
 'to the old and new readers',
 'biographical note',
 "author's taunt",
 'dedication to king of Spain',
 'to Papists',
 'prefaces',
 'verse summary',
 'exhortation to reformation of church',
 'prescriptions']
FRONT_DIV2S = ['encomium',
 'poem',
 'query',
 'letter',
 'question',
 'note',
 'distich',
 'dedication',
 'epigram',
 'to the reader',
 'epitaph',
 'errata',
 'preface',
 'confession',
 'commandment',
 'epigraph',
 'statement',
 'dialogue',
 'sonnet',
 'instruction',
 'petition',
 'definition',
 'observation',
 'prologue',
 'account',
 'postscript',
 'answer',
 'song',
 'comment',
 'rule',
 'question and answer',
 'sequence',
 'conclusion',
 'translation',
 'commendation',
 'to the ill-willed reader',
 'introduction',
 'glossary',
 "author's censure",
 "author's answer",
 'reply',
 'request',
 'oath',
 'omissions',
 'statement in verse',
 'author to his book',
 'envoy',
 'response',
 'exhortation',
 'prayer for the translator',
 'note to the reader',
 'testimonial',
 'book to the reader',
 'declaration',
 'punctuation',
 'anagram',
 'explanation',
 'encomia',
 'to the author',
 'marginal errata',
 'text',
 'preface to contents',
 "Fecknam's dedication",
 'errata in book',
 'letters',
 'poem to the reader',
 'author defiantly to the reader',
 'to the candid reader',
 'addendum']

#BODY
BODY_DIV1S = ['to the reader',
 'dedication',
 'preface',
 'errata',
 'Epistle',
 'author to the reader',
 "author's preface",
 'printer to the reader',
 'note to the reader',
 "author's note",
 'encomia',
 'notice to the reader',
 "translator's note",
 'author to book',
 'to the author',
 'translator to the reader',
 "editor's note",
 "author's prologue",
 'to the translator',
 'prefatory letter',
 "translator's preface",
 'to the Christian reader',
 'poem to the reader',
 'author to ballad makers',
 'author to poets',
 'author to publisher',
 'front matter',
 "author's apology",
 "author's dedication",
 'glossary of terms',
 'instructions to the reader',
 "publisher's note",
 'to scholars',
 'prefatory address',
 'rebuttal to front matter',
 'admonition to the reader',
 'author to brother',
 'letter to the reader',
 "printer's note",
 'introduction to the reader',
 'book to the reader',
 'to the detractors',
 'author to printer',
 'publisher to the reader',
 'author to the work',
 "translator's conclusion",
 'poem to pupils',
 "translator's epilogue",
 'to Catholic readers',
 "translator's introduction",
 'author to colleagues',
 'to his book',
 "author's statement"]
BODY_DIV2S = ['dedication',
 'to the reader',
 'encomium',
 'exhortation',
 "translator's note",
 "author's preface",
 "translator's preface",
 'errata',
 'printer to the reader',
 'dedicatory poem',
 'translator to the reader',
 'encomia',
 'general preface',
 'prefatory poem',
 'author to the reader',
 'acrostic introductory encomium',
 'quoted front matter',
 'preface on prayer',
 'to the Christian reader',
 'epistle to the reader',
 "editor's note",
 'to printer',
 'notice to the reader',
 "editor's envoy",
 'addition by printer',
 "Bishop's preface",
 "author's introduction",
 "author's note",
 'epigrams, epitaphs, and anagrams',
 "publisher's note",
 'warning about dosages']


# BACK
BACK_DIV1S = ['errata',
 'poem',
 'to the reader',
 'epilogue',
 'summary',
 'dedication',
 'prayer',
 'letter',
 'encomium',
 'postscript',
 'note',
 'conclusion',
 'glossary',
 "publisher's advertisement",
 'addendum',
 'printer to the reader',
 'notice',
 'translator to the reader',
 "author's note",
 'afterword',
 'elegy',
 'author to the reader',
 'envoy',
 'text',
 'sonnet',
 'erratum',
 'notes',
 'note to the reader',
 'corrigenda',
 'advertisement',
 'author to book',
 "translator's note",
 "publisher's note",
 'abstract',
 'riddle',
 'dialogue',
 'advice',
 'acrostic',
 'encomia',
 'poem to the reader',
 "publisher's postscript",
 'argument',
 'biography',
 "printer's note",
 'preface',
 'epigram',
 'attestation',
 'advert for medicines',
 'by the same author',
 'apology',
 'admonition',
 'compositor to the reader',
 'postscript to the reader',
 'advert for mathematical instruments',
 'supplement',
 'conclusion to the reader',
 'observation',
 'bibliographic note',
 'notice regarding deferred publication',
 'appreciation',
 'directions for using diagrams',
 'from author',
 'prologue',
 'royal edict on price',
 'index of manuscript sources',
 'printer to author poem',
 'advert for oils',
 'note on the text',
 'addenda and errata',
 'instructions pertaining to psalms',
 'publisher to the reader',
 'notice to the reader',
 'sentence',
 'apology in lieu of errata',
 'response',
 "editor's note",
 'book to the reader',
 "translator's prayer",
 'poems of acknowledgment',
 'treatise',
 "translator's afterword",
 "printer's afterword",
 'correspondence',
 'main index',
 'challenge',
 'summary of part two (forthcoming)',
 'apology and dedication',
 'summary of errors by Doctor Whitgifies',
 'index to further reading',
 "printer's notice",
 'to doubtful reader',
 'instructions pertaining to tunes',
 'introduction to numeracy',
 'letter to author',
 'acrostic encomium',
 'translator to king',
 'friend of author to the reader',
 'introduction',
 'poem in honor of Gustavus',
 'to the Christian reader',
 'laudatory poem',
 'dedication and acknowledgements',
 'collection',
 'authors to the reader',
 'royal thanks',
 "author's addendum",
 'to the academic reader',
 'list of mistranslations',
 "translator's conclusion",
 'epigrams',
 'Bishop to the reader',
 'commentary',
 'explication',
 'synopsis',
 'corrigenda and addenda',
 'index to addendum',
 'authorities',
 'Britain to author',
 "Caxton's epilogue",
 'apology to the reader',
 'notice regarding errata',
 'questions',
 'invocation',
 'caveat to the reader',
 'dedicatory poem',
 'acknowledgment',
 'letters and poems',
 'acknowledgement',
 'record',
 'applausus',
 "printer's additions",
 'to the author']
BACK_DIV2S = ['letter',
 'sonnet',
 'prayer',
 'poem',
 'dialogue',
 'errata',
 'addendum',
 'encomium',
 'addition',
 'note',
 'to the reader',
 'dedication',
 'preface',
 'dedicatory poem',
 'errata for the preface',
 'errata (marginal notes)',
 'conclusion',
 'introduction',
 'response',
 'envoy',
 'textual errata',
 'advice']

# OTHER
pluraltypes = ['encomia', 'epigraphs', 'quotations', 'poems', 'observations', 'epitaphs', 'instructions', 'additions', 'dedications', 'arguments', 'articles', 'sonnets', 'epigrams, etc.', 'letters', 'endorsements', 'songs', 'prefaces', 'prayers', 'epigrams' 'letters and poems']
tabletypes = ['errata']


# Define 'Describe' function

def describe(preface):
    #this function takes a preface xml, identifies relevant data and analysis
    #and writes it to an excel file
    
    #head text
    if preface.find('HEAD'):
        head = preface.find('HEAD')
        sheet['G' + str(row)].value = head.text  

    # whole text
    whole_text = preface.text
    whole_text_stripped = whole_text.strip()
    sheet['H' + str(row)].value = whole_text_stripped[:32767] #text sliced to excel's maximum character count

    # word count of whole text
    wordcount = len(whole_text_stripped.split())
    sheet['F' + str(row)].value = wordcount

    #currently skips language detection because limit on TextBlob reached before the time.sleep() pause was added

    # #get language of whole text and assign to dict
    # if len(whole_text_stripped) > 3:
    #     b = TextBlob(whole_text_stripped)
    #     time.sleep(0.2)             #pauses to prevent 'HTTP Error 429: Too Many Requests'
    #     language = b.detect_language()
    #     sheet['S' + str(row)].value = language


    # signed text
    if preface.find('SIGNED'):
        signed = preface.find('SIGNED')
        sheet['I' + str(row)].value = signed.text

    # dateline text
    if preface.find('DATELINE'):
        dateline = preface.find('DATELINE')
        sheet['J' + str(row)].value = dateline.text

    # type
    sheet['E' + str(row)].value = type

    #check if prose/verse by absence of <P>
    para = preface.find_all('P')
    verse = preface.find_all('L')

    if wordcount == 0:
        form = 'N/A'
    elif len(para) > 0 and len(verse) == 0:
        form = 'Prose'
    elif len(para) == 0 and len(verse) > 0:
        form = 'Verse'
    elif len(para) > 0 and len(verse) > 0:
        form = 'Prose/Verse'
    else:
        form = 'Other'

    sheet['Q' + str(row)].value = form

    #filename
    sheet['C' + str(row)].value = file

    #position
    if position == 'front':
        sheet['D' + str(row)].value = 'front' + str(front_count)
    if position == 'body':
        sheet['D' + str(row)].value = 'body' + str(body_count)
    elif position == 'back':
        sheet['D' + str(row)].value = 'back' + str(back_count)

    #preface ID
    sheet['B' + str(row)].value = ID

    #Edition
    sheet['A' + str(row)].value = ESTC

    #add preface ID to edition entry
    if editionsheet['O' + str(filecount + 2)].value:
        editionsheet['O' + str(filecount + 2)].value = editionsheet['O' + str(filecount + 2)].value + ID + '; '
    else:
        editionsheet['O' + str(filecount + 2)].value = ID + '; '

    # create year folder for xml output if it doesn't exist    
    if not Path('Output', year).is_dir:
        Path('Output', year).mkdir()

    # #save preface as separate xml file
    with open(str(Path('Output') / year / ID / '.xml'), 'w') as f:
        f.write(str(preface))


# Set output spreadsheet

wb = openpyxl.load_workbook('TTR Jan.xlsx')
sheet = wb['Preface']
editionsheet = wb['Edition']


filecount = 0

# ***** define directory here ***********
# Either do all at once (as it is), or in batches by specifying the decade folders after 'Input' in the line below (e.g. 'Input\\1473–1480', etc.)

for root, dirs, files in os.walk('Input'):
    print(root)
    for file in files:
        if file.endswith('.xml'): #added to avoid reading .DS_Store and other file types
            print(file)
            filecount += 1
            
            #open file as BS xml
            with open(str(directory / root / file), 'r' ,encoding='utf-8') as currentfile:

                year = str(root)[-4:]

                content = currentfile.readlines()
                content = ''.join(content)
                bs_content = bs(content, 'xml')

                #grab header
                header = bs_content.find('HEADER')
                EEBO = bs_content.find('EEBO')

                #get ESTC no
                IDNO_stc_list = header.find_all('IDNO', {'TYPE' : 'stc'}) #this is the primary source of the ESTC number
                STC_T_C = EEBO.find('STC', {'T' : 'C'}) #but if the above tag doesn't exist in the file, this should have it

                if len(IDNO_stc_list) >= 2:
                    ESTC_text = IDNO_stc_list[1].text
                    ESTC = ESTC_text[5:]
                elif STC_T_C != None:
                    ESTC = STC_T_C.text
                else:
                    ESTC = 'Unknown_' + file #if the file doesnt have either <IDNO T:stc> or <STC T:C>, then the ESTC is this, and should be corrected manually

                #put ESTC number and file name into Edition table

                editionsheet['A' + str(filecount + 2)].value = ESTC
                editionsheet['N' + str(filecount + 2)].value = str(file)

                #start counters
                front_count = 0
                body_count = 0
                back_count = 0

                #FRONTS
                fronts = bs_content.find_all('FRONT') #returns a list
                position = 'front'

                for front in fronts:
                    front_count = front_count + 1

                    divs = front.find_all('DIV1') #gives a list

                    #go through each div in current front
                    for div in divs:
                        type = div.get('TYPE')  
                            
                        if type in pluraltypes:
                            div2s = div.find_all('DIV2')

                            if len(div2s) > 0:
                                for div2 in div2s:
                                    if div2.get('TYPE') in FRONT_DIV1S:
                                        
                                        type = div2.get('TYPE')
                                        ID = ESTC + '_Fr' + str(fronts.index(front) + 1) + '_' + str(divs.index(div) + 1) + '_' + str(div2s.index(div2) + 1)
                                        row = sheet.max_row + 1
                                        describe(div2)

                            else:
                                
                                ID = ESTC + '_Fr' + str(fronts.index(front) + 1) + '_' + str(divs.index(div) + 1)
                                row = sheet.max_row + 1
                                describe(div)


                        elif type in tabletypes: # this loop is to filter out divs that contain tables, unless they also contain poetry/prose

                            Ps_and_Ls = div.find_all(['P', 'L']) #creates a list of all internal P and L tags

                            for item in Ps_and_Ls:
                                if not item.find_all('TABLE'): #goes through list and finds any P/L tags that DONT contain a table (i.e. the prose/poetry ones we are interested in)
                                
                                    ID = ESTC + '_Fr' + str(fronts.index(front) + 1) + '_' + str(divs.index(div) + 1)
                                    row = sheet.max_row + 1
                                    describe(div)
                                    break #breaks after it has found 1, as otherwise would 'describe' the div every time for each tag not containing a table
                                    
                        elif type in FRONT_DIV1S:
                            
                            ID = ESTC + '_Fr' + str(fronts.index(front) + 1) + '_' + str(divs.index(div) + 1)
                            row = sheet.max_row + 1 
                            describe(div)
                        
                        else:                                   #if div1 type in excludedtypes, check internal div2s
                            div2s = div.find_all('DIV2')
                            for div2 in div2s:
                                div1type = type
                                type = div2.get('TYPE')
                                N = div2.get('N')

                                if 'index' in div1type and N:
                                    continue

                                elif type in FRONT_DIV2S:
                                    
                                    ID = ESTC + '_Fr' + str(fronts.index(front) + 1) + '_' + str(divs.index(div) + 1) + '_' + str(div2s.index(div2) + 1)
                                    row = sheet.max_row + 1
                                    describe(div2)
    

                #BODYS
                bodies = bs_content.find_all('BODY') #returns a list

                for body in bodies:
                    position = 'body'
                    body_count = body_count + 1

                    divs = body.find_all('DIV1') #gives a list

                    #go through each div in current body
                    for div in divs:
                        type = div.get('TYPE')

                        if type in BODY_DIV1S:

                            ID = ESTC + '_Bo' + str(bodies.index(body) + 1) + '_' + str(divs.index(div) + 1)
                            row = sheet.max_row + 1
                            describe(div)

                        else:                                   #if div1 type not in bodytypes, check internal div2s
                            div2s = div.find_all('DIV2')
                            for div2 in div2s:
                                type = div2.get('TYPE')

                                if type in BODY_DIV2S:

                                    ID = ESTC + '_Bo' + str(bodies.index(body) + 1) + '_' + str(divs.index(div) + 1) + '_' + str(div2s.index(div2) + 1)
                                    row = sheet.max_row + 1
                                    describe(div2)

                #BACKS
                backs = bs_content.find_all('BACK') # returns a list

                for back in backs:
                    position = 'back'
                    back_count = back_count + 1

                    divs = back.find_all('DIV1') #gives a list

                    #go through each div in current back
                    for div in divs:
                        type = div.get('TYPE')  
                            
                        # check if likely to contain relevant DIV2s    
                        if type in pluraltypes:
                            div2s = div.find_all('DIV2')
                        
                            if len(div2s) > 0:
                                for div2 in div2s:
                                    div2_type = div2.get('TYPE')
                                    if type in BACK_DIV2S:

                                        type = div2_type

                                        ID = ESTC + '_Ba' + str(backs.index(back) + 1) + '_' + str(divs.index(div) + 1) + '_' + str(div2s.index(div2) + 1)
                                        row = sheet.max_row + 1
                                        describe(div2)

                            else:
                                ID = ESTC + '_Ba' + str(backs.index(back) + 1) + '_' + str(divs.index(div) + 1)
                                row = sheet.max_row + 1
                                describe(div)
                                    
                        elif type in tabletypes: # this loop is to filter out divs that contain tables, unless they also contain poetry/prose

                            Ps_and_Ls = div.find_all(['P', 'L']) #creates a list of all internal P and L tags

                            for item in Ps_and_Ls:
                                if not item.find_all('TABLE'): #goes through list and finds any P/L tags that DONT contain a table (i.e. the prose/poetry ones we are interested in)
                                
                                    ID = ESTC + '_Ba' + str(backs.index(back) + 1) + '_' + str(divs.index(div) + 1)
                                    row = sheet.max_row + 1
                                    describe(div)
                                    break #breaks after it has found 1, as otherwise would 'describe' the div every time for each tag not containing a table
                                                    

                        elif type in BACK_DIV1S:
                            
                            ID = ESTC + '_Ba' + str(backs.index(back) + 1) + '_' + str(divs.index(div) + 1)
                            row = sheet.max_row + 1
                            describe(div)


                        else:                                   #if div1 type in excludedtypes, check internal div2s
                            div2s = div.find_all('DIV2')
                            for div2 in div2s:
                                div1type = type
                                type = div2.get('TYPE')
                                N = div2.get('N')

                                if 'index' in div1type and N:
                                    continue

                                elif type in BACK_DIV2S: 

                                    ID = ESTC + '_Ba' + str(backs.index(back) + 1) + '_' + str(divs.index(div) + 1) + '_' + str(div2s.index(div2) + 1)
                                    row = sheet.max_row + 1
                                    describe(div2)

    wb.save('TTR Jan.xlsx')


print("--- %s seconds ---" % (time.time() - start_time))